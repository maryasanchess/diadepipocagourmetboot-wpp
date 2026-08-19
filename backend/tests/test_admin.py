"""Testes dos comandos administrativos e do relatório mensal."""

from datetime import datetime

from app import admin, dispatcher, models
from app.config import settings


def test_normaliza_telefone_sem_ddi():
    assert admin._normalizar_telefone("11988887777") == "5511988887777"


def test_normaliza_telefone_com_ddi_mantem():
    assert admin._normalizar_telefone("5511988887777") == "5511988887777"


def test_eh_admin_aceita_multiplos_numeros(monkeypatch):
    monkeypatch.setattr(settings, "admin_phone_number", "5511999990001,11999990002")
    assert admin.eh_admin("5511999990001") is True
    assert admin.eh_admin("5511999990002") is True  # mesmo sem DDI no .env
    assert admin.eh_admin("5511111111111") is False


def test_relatorio_agrega_por_sabor_e_tamanho(db):
    cliente = models.Cliente(telefone="5511999990000")
    db.add(cliente)
    db.commit()
    db.refresh(cliente)

    pedido = models.Pedido(
        cliente_id=cliente.id,
        tipo_entrega="retirada",
        status=models.StatusPedido.recebido,
    )
    pedido.itens.append(models.ItemPedido(sabor="Nutella", tamanho_g=100, quantidade=3, preco_unitario=18.0))
    pedido.itens.append(models.ItemPedido(sabor="Kinder Bueno", tamanho_g=100, quantidade=1, preco_unitario=22.0))
    db.add(pedido)

    # pedido cancelado não deve entrar no relatório
    cancelado = models.Pedido(
        cliente_id=cliente.id,
        tipo_entrega="retirada",
        status=models.StatusPedido.cancelado,
    )
    cancelado.itens.append(models.ItemPedido(sabor="Ninho", tamanho_g=80, quantidade=5, preco_unitario=12.0))
    db.add(cancelado)
    db.commit()

    from app.relatorio import gerar_relatorio_mensal
    from openpyxl import load_workbook

    agora = datetime.now()
    caminho = gerar_relatorio_mensal(db, agora.year, agora.month)
    try:
        wb = load_workbook(caminho)
        ws = wb.active
        linhas = list(ws.iter_rows(values_only=True))

        assert ("Kinder Bueno", 100, 1, 22) in linhas
        assert ("Nutella", 100, 3, 54) in linhas
        assert ("Total de pedidos", 1, None, None) in linhas  # só o não cancelado
        assert ("Faturamento total", 76, None, None) in linhas
    finally:
        caminho.unlink(missing_ok=True)


def test_dispatcher_admin_com_comando_reconhecido(db, monkeypatch):
    monkeypatch.setattr(settings, "admin_phone_number", "5511999990001")
    resposta = dispatcher.processar_mensagem_recebida(db, "5511999990001", "relatorio")

    assert "relatório" in resposta.texto.lower()
    assert resposta.anexo is not None
    assert resposta.anexo.exists()
    assert resposta.anexo.suffix == ".xlsx"

    resposta.anexo.unlink()


def test_dispatcher_admin_sem_comando_reconhecido_vira_pedido_normal(db, monkeypatch):
    monkeypatch.setattr(settings, "admin_phone_number", "5511999990001")
    resposta = dispatcher.processar_mensagem_recebida(db, "5511999990001", "oi")
    assert "bem-vindo" in resposta.texto.lower()
    assert resposta.anexo is None
