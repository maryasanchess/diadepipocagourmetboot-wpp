"""
Geração da planilha mensal: quantidade vendida e faturamento por sabor e
tamanho, para um mês específico (o mês atual, por padrão).

O arquivo gerado fica em backend/relatorios/ — nunca é commitado (contém
dados reais de vendas), ver .gitignore.
"""

import calendar
from datetime import datetime
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from sqlalchemy.orm import Session

from . import models
from .config import BASE_DIR

RELATORIOS_DIR = BASE_DIR / "relatorios"
RELATORIOS_DIR.mkdir(parents=True, exist_ok=True)

COR_CABECALHO = "C87F0A"


def _intervalo_do_mes(ano: int, mes: int) -> tuple[datetime, datetime]:
    inicio = datetime(ano, mes, 1)
    ultimo_dia = calendar.monthrange(ano, mes)[1]
    fim = datetime(ano, mes, ultimo_dia, 23, 59, 59)
    return inicio, fim


def gerar_relatorio_mensal(db: Session, ano: int | None = None, mes: int | None = None) -> Path:
    """Gera o .xlsx do mês informado (ou o mês atual) e retorna o caminho do arquivo."""
    agora = datetime.now()
    ano = ano or agora.year
    mes = mes or agora.month
    inicio, fim = _intervalo_do_mes(ano, mes)

    pedidos = (
        db.query(models.Pedido)
        .filter(models.Pedido.criado_em >= inicio, models.Pedido.criado_em <= fim)
        .filter(models.Pedido.status != models.StatusPedido.cancelado)
        .all()
    )

    resumo: dict[tuple[str, int], dict] = {}
    faturamento_total = 0.0
    for pedido in pedidos:
        for item in pedido.itens:
            chave = (item.sabor, item.tamanho_g)
            linha = resumo.setdefault(chave, {"quantidade": 0, "faturamento": 0.0})
            subtotal = item.quantidade * item.preco_unitario
            linha["quantidade"] += item.quantidade
            linha["faturamento"] += subtotal
            faturamento_total += subtotal

    wb = Workbook()
    ws = wb.active
    ws.title = f"{mes:02d}-{ano}"

    cabecalho = ["sabor", "tamanho_g", "quantidade", "faturamento"]
    ws.append(cabecalho)
    for celula in ws[1]:
        celula.font = Font(bold=True, color="FFFFFF")
        celula.fill = PatternFill(start_color=COR_CABECALHO, end_color=COR_CABECALHO, fill_type="solid")

    for (sabor, tamanho_g), dados in sorted(resumo.items(), key=lambda item: (item[0][0], item[0][1])):
        ws.append([sabor, tamanho_g, dados["quantidade"], round(dados["faturamento"], 2)])

    ws.append([])
    ws.append(["Total de pedidos", len(pedidos)])
    ws.append(["Faturamento total", round(faturamento_total, 2)])

    for coluna in ws.columns:
        valores = [str(celula.value) for celula in coluna if celula.value is not None]
        largura = max((len(v) for v in valores), default=8) + 2
        ws.column_dimensions[coluna[0].column_letter].width = largura

    caminho = RELATORIOS_DIR / f"relatorio_{ano}-{mes:02d}.xlsx"
    wb.save(caminho)
    return caminho
